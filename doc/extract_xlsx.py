import openpyxl
import io
import sys

path = r"E:/antigravity/rag/doc/한국지식재산보호원 AI 영업비밀_기능요구사항검토20260427_1_로이드케이체크.xlsx"
out_path = r"E:/antigravity/rag/doc/_xlsx_dump.txt"

wb = openpyxl.load_workbook(path, data_only=True)

lines = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    lines.append(f"\n========== SHEET: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ==========")
    for row in ws.iter_rows(values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        cells = ["" if v is None else str(v).replace("\n", " \\n ").strip() for v in row]
        lines.append(" | ".join(cells))

with io.open(out_path, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print("done:", out_path)
