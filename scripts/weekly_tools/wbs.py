# -*- coding: utf-8 -*-
"""
wbs.py — [지재원] AI 영업비밀관리시스템 구축 WBS.xlsx 읽기/쓰기 컴포넌트

간트 구조(고정):
  - 축: 열 G(=7)부터 하루 1열, 시작일 2026-06-02(화), ~2027-01-04.
  - 행1=월 띠(병합), 행2=주차(병합, 월 내부 nest), 행3=일자, 행4=요일/라벨헤더.
  - 작업 행 = 5..(마지막 작업행). 틀고정 G5.
  - 라벨열: B=구분, C=작업, D=기능상세, E=산출물, F=상태.
  - 막대 색: 완료 FF3F7D58 · 마일스톤 FFFFC000 · 예정 FF9EADCC/FF6E8F72/FF8064A2.
  - 휴무 회색 FFD9D9D9 · 공휴일 머리글 빨강 FFC00000.

주요 사용법:
    w = WBSGantt(path)
    for t in w.tasks(): print(t.row, t.gubun, t.jakup, t.status)
    w.set_status(29, "완료")
    w.set_bar(29, date(2026,7,6), date(2026,7,24), WBSGantt.C_DONE)
    w.reapply_formatting()   # 달력 띠·휴무 음영·테두리·폰트 재도색(idempotent)
    w.save()                 # 원본 덮어쓰기 (또는 save(other_path))
"""
import datetime
from copy import copy
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from . import kr_calendar as cal

GRID_START_COL = 7                       # 'G'
GRID_START_DATE = datetime.date(2026, 6, 2)
FIRST_TASK_ROW = 5


@dataclass
class Task:
    row: int
    gubun: str          # 구분 (대공정)
    jakup: str          # 작업
    detail: str         # 기능상세
    output: str         # 산출물
    status: str         # 상태 / 목표시점
    bar_start: datetime.date
    bar_end: datetime.date
    bar_color: str


def _hexfill(cell):
    f = cell.fill
    if f and f.patternType and f.fgColor and str(f.fgColor.rgb) != "00000000":
        return str(f.fgColor.rgb)
    return None


class WBSGantt:
    # 색상 상수
    C_DONE = "FF3F7D58"       # 완료(작업/1차)
    C_MILESTONE = "FFFFC000"  # 마일스톤
    C_PLAN_BLUE = "FF9EADCC"  # 예정 — 데이터검수·최종검증·완료보고
    C_PLAN_OLIVE = "FF6E8F72" # 예정 — 배포준비·시스템연계
    C_PLAN_PURPLE = "FF8064A2"# 예정 — KL API 협의·베타·감리
    GREY = "FFD9D9D9"         # 휴무(주말·공휴일)
    RED = "FFC00000"          # 공휴일 머리글
    HDR_BLUE = "FF055C9D"     # 월/주차 헤더 배경
    FONT = "Malgun Gothic"

    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb.active
        self.start_col = GRID_START_COL
        self.start_date = GRID_START_DATE
        self._verify_calendar_anchor()

    # ---- 좌표/날짜 매핑 ----
    def col_of(self, d):
        return self.start_col + (d - self.start_date).days

    def date_of(self, c):
        return self.start_date + datetime.timedelta(days=c - self.start_col)

    @property
    def last_col(self):
        c = self.ws.max_column
        while c >= self.start_col and self.ws.cell(row=3, column=c).value is None:
            c -= 1
        return c

    @property
    def last_task_row(self):
        """작업 블록은 행5부터 '연속'이다. 첫 빈 행 직전까지가 작업 범위
        (아래쪽 범례 블록을 작업으로 오인하지 않도록 위에서부터 스캔)."""
        last = FIRST_TASK_ROW
        r = FIRST_TASK_ROW
        while r <= self.ws.max_row:
            if any(self.ws.cell(row=r, column=col).value not in (None, "")
                   for col in (2, 3, 4, 5, 6)):   # B..F (E=산출물 포함)
                last = r
                r += 1
            else:
                break
        return last

    def _verify_calendar_anchor(self):
        """행3/행4가 start_date 가정과 일치하는지 확인(어긋나면 예외)."""
        d3 = self.ws.cell(row=3, column=self.start_col).value
        w4 = self.ws.cell(row=4, column=self.start_col).value
        ok = (d3 is not None and str(d3).strip().isdigit()
              and int(d3) == self.start_date.day
              and w4 == cal.kweekday(self.start_date))
        if not ok:
            raise ValueError(
                f"달력 기준점 불일치: G3={d3!r}, G4={w4!r} vs 가정 {self.start_date} "
                f"({cal.kweekday(self.start_date)}). 파일 축이 바뀌었을 수 있음.")

    # ---- 읽기 ----
    def tasks(self):
        out = []
        lastB = lastC = ""
        for r in range(FIRST_TASK_ROW, self.last_task_row + 1):
            B = self.ws.cell(row=r, column=2).value
            C = self.ws.cell(row=r, column=3).value
            if B:
                lastB = B
            if C:
                lastC = C
            bs, be, bc = self._bar_span(r)
            out.append(Task(
                row=r,
                gubun=(B or lastB or ""),
                jakup=(C or lastC or ""),
                detail=self.ws.cell(row=r, column=4).value or "",
                output=self.ws.cell(row=r, column=5).value or "",
                status=self.ws.cell(row=r, column=6).value or "",
                bar_start=bs, bar_end=be, bar_color=bc,
            ))
        return out

    def _bar_span(self, r):
        cols = [c for c in range(self.start_col, self.last_col + 1)
                if _hexfill(self.ws.cell(row=r, column=c)) not in (None, self.GREY)]
        if not cols:
            return None, None, None
        color = _hexfill(self.ws.cell(row=r, column=cols[0]))
        return self.date_of(min(cols)), self.date_of(max(cols)), color

    # ---- 쓰기: 상태 ----
    def set_status(self, row, text):
        self.ws.cell(row=row, column=6).value = text

    # ---- 쓰기: 막대 ----
    def clear_bar(self, row):
        """행의 막대를 모두 제거(휴무 회색은 reapply_formatting이 복구)."""
        for c in range(self.start_col, self.last_col + 1):
            if _hexfill(self.ws.cell(row=row, column=c)) not in (None, self.GREY):
                self.ws.cell(row=row, column=c).fill = PatternFill(fill_type=None)

    def set_bar(self, row, start, end, color, workdays_only=True):
        """start~end 구간에 막대를 칠함. workdays_only면 근무일(주말·공휴일 제외)만.
        기존 막대는 먼저 지움. reapply_formatting()으로 휴무 음영을 복구할 것."""
        self.clear_bar(row)
        fill = PatternFill(fill_type="solid", fgColor=color)
        d = start
        while d <= end:
            c = self.col_of(d)
            if self.start_col <= c <= self.ws.max_column:
                if (not workdays_only) or cal.is_workday(d):
                    self.ws.cell(row=row, column=c).fill = fill
            d += datetime.timedelta(days=1)

    def set_milestone(self, row, on_date):
        self.clear_bar(row)
        c = self.col_of(on_date)
        if self.start_col <= c <= self.ws.max_column:   # 그리드 밖이면 무시(set_bar과 동일)
            self.ws.cell(row=row, column=c).fill = \
                PatternFill(fill_type="solid", fgColor=self.C_MILESTONE)

    # ---- 재도색(idempotent): 달력 띠 · 휴무 음영 · 테두리 · 폰트 ----
    def reapply_formatting(self):
        self._paint_calendar_bands()
        self._paint_offdays()
        self._unify_fonts()
        self._unify_borders()

    def _hdr_style(self, r, c, val=None):
        cell = self.ws.cell(row=r, column=c)
        cell.fill = PatternFill(fill_type="solid", fgColor=self.HDR_BLUE)
        cell.font = Font(name=self.FONT, size=10, bold=True, color="FFFFFFFF")
        med = Side(style="medium", color="FF000000")
        cell.border = Border(left=med, right=med, top=med, bottom=med)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if val is not None:
            cell.value = val

    def _paint_calendar_bands(self):
        """행1 월띠(정확한 월경계) + 행2 주차(월 내부 nest, <3일 슬리버 흡수)."""
        ws = self.ws
        lc = self.last_col
        # 간트 열(G+)에 걸치는 행1/2 병합은 모두 해제(경계를 넘어오는 병합 포함)
        for m in [mm for mm in list(ws.merged_cells.ranges)
                  if mm.min_row in (1, 2) and mm.max_col >= self.start_col]:
            ws.unmerge_cells(str(m))
        for r in (1, 2):
            for c in range(self.start_col, lc + 1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):   # 잔존 병합셀 .value 쓰기 방지
                    cell.value = None
        cols = list(range(self.start_col, lc + 1))
        # 월 그룹
        mgroups = []
        for c in cols:
            d = self.date_of(c)
            key = (d.year, d.month)
            if mgroups and mgroups[-1][0] == key:
                mgroups[-1][1].append(c)
            else:
                mgroups.append([key, [c]])
        for (yr, mo), mcols in mgroups:
            for c in mcols:
                self._hdr_style(1, c)
            ws.cell(row=1, column=mcols[0]).value = f"{mo}월"
            if len(mcols) > 1:
                ws.merge_cells(start_row=1, start_column=mcols[0],
                               end_row=1, end_column=mcols[-1])
            # 주차: 글로벌 7일블록으로 나눈 뒤 <3일은 이웃에 흡수
            subs = []
            for c in mcols:
                wid = (c - self.start_col) // 7
                if subs and subs[-1][0] == wid:
                    subs[-1][1].append(c)
                else:
                    subs.append([wid, [c]])
            groups = [s[1] for s in subs]
            changed = True
            while changed and len(groups) > 1:
                changed = False
                for i, g in enumerate(groups):
                    if len(g) < 3:
                        if i > 0:
                            groups[i - 1] = groups[i - 1] + g
                        else:
                            groups[i + 1] = g + groups[i + 1]
                        groups.pop(i)
                        changed = True
                        break
            for n, g in enumerate(groups, 1):
                for c in g:
                    self._hdr_style(2, c)
                ws.cell(row=2, column=g[0]).value = f"{n} 주차"
                if len(g) > 1:
                    ws.merge_cells(start_row=2, start_column=g[0],
                                   end_row=2, end_column=g[-1])

    def _paint_offdays(self):
        """주말·공휴일: 본문(5..마지막행) 회색, 공휴일은 머리글(행3·4) 빨강.
        막대가 있는 칸은 덮어쓰지 않음(막대 유지)."""
        ws = self.ws
        lr = self.last_task_row
        grey = PatternFill(fill_type="solid", fgColor=self.GREY)
        red = PatternFill(fill_type="solid", fgColor=self.RED)
        BARS = {self.C_DONE, self.C_MILESTONE, self.C_PLAN_BLUE,
                self.C_PLAN_OLIVE, self.C_PLAN_PURPLE}
        for c in range(self.start_col, self.last_col + 1):
            d = self.date_of(c)
            off = cal.is_weekend(d) or cal.is_holiday(d)
            if off:
                for r in range(FIRST_TASK_ROW, lr + 1):
                    if _hexfill(ws.cell(row=r, column=c)) not in BARS:
                        ws.cell(row=r, column=c).fill = grey
            if cal.is_holiday(d):
                for r in (3, 4):
                    ws.cell(row=r, column=c).fill = red

    def _unify_fonts(self):
        """가시 텍스트 전부 Malgun Gothic(크기/굵기/색 보존)."""
        for r in range(1, self.last_task_row + 8):
            for c in range(1, self.ws.max_column + 1):
                cell = self.ws.cell(row=r, column=c)
                f = cell.font
                if f and f.name and f.name != self.FONT:
                    nf = copy(f)
                    nf.name = self.FONT
                    cell.font = nf

    def _unify_borders(self):
        """라벨표 B4:F(마지막행): 굵은선=외곽+열세로선+구분 경계 / 일반선=구분 내부 행선."""
        ws = self.ws
        lr = self.last_task_row
        med = Side(style="medium", color="FF000000")
        thin = Side(style="thin", color="FF000000")
        # 구분(B열) 그룹 시작행 = B에 값이 있는 행 (+ 마일스톤/헤더)
        starts = {FIRST_TASK_ROW}
        prevB = None
        for r in range(FIRST_TASK_ROW, lr + 1):
            b = ws.cell(row=r, column=2).value
            if b and b != prevB:
                starts.add(r)
            if b:
                prevB = b
        target = [str(m) for m in ws.merged_cells.ranges
                  if 4 <= m.min_row and m.max_row <= lr and 2 <= m.min_col and m.max_col <= 6]
        for m in target:
            ws.unmerge_cells(m)
        medium_top = set(starts) | {4}
        medium_bottom = {r - 1 for r in starts if r - 1 >= 4} | {lr, 4}
        for r in range(4, lr + 1):
            T = med if r in medium_top else thin
            Bt = med if r in medium_bottom else thin
            for c in range(2, 7):
                ws.cell(row=r, column=c).border = Border(left=med, right=med, top=T, bottom=Bt)
        for m in target:
            ws.merge_cells(m)

    # ---- 저장 ----
    def save(self, path=None):
        self.wb.save(path or self.path)
        return path or self.path
