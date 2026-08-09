#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""골든셋 관리자 검수 워크북(.xlsx) 발행 — 지재원 관리자가 등급을 판정해 돌려주는 기입형 시트.

왜 엑셀인가: 기존 검수 경로는 CSV 시트 3종(golden500 354행·golden50 35행·human_review_queue
80행)과 서명 HTML 이 있었고, 기입 회수는 0/389 였다(signed.csv 40건은 reviewer_id=ai_assist 로
전건 무효). 형식이 문제였던 적은 없으므로 이 도구는 **회수 경로**를 같이 만든다 —
import_golden_review_xlsx.py 가 이 워크북을 받아 golden_signoff.promote_to_locked 로만 승격한다.

이 도구가 하지 않는 것:
  - 정본 datasets/gold_real/classification_gold.jsonl 을 쓰지 않는다(읽기 전용).
  - 배포모델 승격 게이트를 신설하지 않는다. 산출 locked 는 **측정용**이다.
  - 사람 서명을 위조 불가로 만들지 못한다. 대조키는 정렬·행삽입·doc_id 손상 같은
    **사고**를 잡는 정합 검사이고, 악의적 위조에 대한 암호학적 증거가 아니다(README 참조).

검수 우선순위(P1이 가장 가치 높음) — 근거는 실측:
  P1 공표평가셋 & 실문서 : 서명이 붙으면 F1 0.917 의 정답지가 machine-silver → 사람서명 승격
  P2 실문서 희소등급     : public_real 인 S1·S2. 학습셋 미포함분만
  P3 합성 고등급         : TS·S1. 본문이 합성이라 real 평가정답 불가 — 회귀 픽스처·검수 실적
  P4 실문서 S3           : 풍부하고 판정이 쉬움(공개 판례 → 정의상 S3). 정보량 낮음
  P5 학습셋 포함         : 검수해도 평가정답으로 쓸 수 없다(train-on-test). 참고 표시만

주의(2026-08-08 정정): eval_readiness 의 ready 는 **사람 서명 건수**만으로 성립한다
(golden_tiers.is_locked_eval = is_valid_signoff, 2026-08-06 변경). 합성 본문이어도 서명이
붙으면 집계된다 — 실제로 기존 서명 34건(전건 합성)만으로 ready=True 가 된다(실측).
따라서 "readiness 도달 불가"는 더 이상 사실이 아니다. 대신 두 가지가 남는다:
  · real_per_grade.TS 는 여전히 0이고 이 파일로는 채울 수 없다(public_real TS = 0건).
    실 비밀문서는 정의상 공개돼 있지 않으므로 발주처 실문서가 와야 열린다.
  · 합성 본문으로 잰 정확도는 실세계 성능이 아니다(자체 실측: 합성학습→실문서 F1 0.26).
    수치를 인용할 때는 real/synthetic 구성을 함께 밝힌다.

⚠ document_origin 신뢰 한계: source="금융보고서" 229건이 public_real 로 집계되지만 원자료는
합성이다(datasets/raw/manifest.yaml:119 nmixx-fin/synthetic_financial_report_korean →
p2_oss_corpus_builder.py:198 → golden_tiers.py:56 _PUBLIC_REAL_EXACT_SOURCES). 정정하면
public_real 641→412 이고 **실문서 S2 는 48→0** 이 된다. 아래 P2("실문서 희소등급")는 그만큼
과대 집계돼 있다 — 우선순위를 신뢰하기 전에 이 매핑을 먼저 정할 것.

사용:
  python scripts/export_golden_review_xlsx.py --out-dir datasets/golden_review
  python scripts/export_golden_review_xlsx.py --limit 60           # 상위 60건만
  python scripts/export_golden_review_xlsx.py --priority P1 P2 P3   # 특정 그룹만
"""
from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import json
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "src"))

from lloydk.golden_signoff import SIGNOFF_GATE_VERSION  # noqa: E402
from lloydk.golden_tiers import document_origin, is_external_authority, tier_of  # noqa: E402
from lloydk.hygiene import text_hash  # noqa: E402

GOLD = _ROOT / "datasets" / "gold_real" / "classification_gold.jsonl"
# 배포본 v-fe4b386b 의 실학습 코퍼스. 구 train_subset.jsonl(610행)이 아니다 —
# 그 파일로 누출을 재면 실학습 본문 대부분이 검사에 안 보인다(실측: 교집합 202/610).
TRAIN_CORPUS = [
    _ROOT / "datasets" / "labeled_p1_v5_clean" / "train.jsonl",
    _ROOT / "datasets" / "labeled_p1_v5_clean" / "val.jsonl",
    _ROOT / "datasets" / "labeled_p1_v5_clean" / "test.jsonl",
]
# 공표 지표(F1 0.917 · 과분류 0.04)가 정답지로 쓰는 평가셋. 여기에 사람 서명이 붙는 것이
# 가장 가치가 크므로 제외 대상이 아니라 **최우선** 대상이다.
EVAL_SETS = {
    "경화42": _ROOT / "datasets" / "gold_real" / "holdout_eval.hardened.jsonl",
    "clean42": _ROOT / "datasets" / "gold_real" / "holdout_eval.clean.jsonl",
    "프록시109": _ROOT / "datasets" / "gold_real" / "holdout_eval.jsonl",
}

GRADES = ("TS", "S1", "S2", "S3")
HOLD = "보류"
CHOICES = GRADES + (HOLD,)

# LABELING_GUIDE.md:85-92 (정본). 문구를 창작하지 않고 그대로 옮긴다.
GRADE_CRITERIA = [
    ("TS", "유출 시 회사 존립 위협 / 국가 안보 영향", "미공개 핵심 기술, 진행 중 M&A 가격, 루트 CA 키"),
    ("S1", "유출 시 경쟁우위 상실 / 영업비밀 침해", "공정 노하우, 원가 구조, 고객 DB"),
    ("S2", "외부 공개 시 협상력 약화 / 이미지 손상", "미확정 사업계획, 내부 검토 초안"),
    ("S3", "공개 예정이거나 이미 공개된 정보", "보도자료, 채용공고, 공시"),
]
FNR_RULE = "모호하면 한 단계 위 등급으로 판정합니다 (미탐 회피 원칙 — LABELING_GUIDE.md:92)."

MAX_CELL = 32_000  # openpyxl 은 32,767 초과를 무경고 절단한다 → 그 앞에서 우리가 막는다
# 상류 파이프라인이 본문을 6,000자에서 자른 흔적(실측 116건이 정확히 6,000자). 판결문 뒷부분이
# 잘린 상태로 등급을 서명받으면 비공지성·경제적유용성 판단 자체가 오염되므로 원문을 복원한다.
TRUNCATION_MARK = 6000
OSS_CORPUS = _ROOT / "datasets" / "oss_corpus"

COLUMNS = [
    # (헤더, 폭, 편집가능)
    ("순번", 6, False),
    ("우선순위", 22, False),
    ("doc_id", 19, False),
    ("대조키", 14, False),
    ("출처", 13, False),
    ("도메인", 12, False),
    ("문자수", 8, False),
    ("본문상태", 20, False),
    ("기계등급(참고)", 13, False),
    ("라벨출처", 20, False),
    ("근거권위", 24, False),
    ("본문파일", 26, False),
    ("본문(전문)", 100, False),
    ("판정등급", 11, True),
    ("판단근거", 42, True),
    ("비고", 24, True),
]
EDIT_HEADERS = [h for h, _, e in COLUMNS if e]


def _read_jsonl(path: Path) -> list[dict]:
    out: list[dict] = []
    if not path.exists():
        return out
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                try:
                    out.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    return out


def _hashes(path: Path) -> set[str]:
    return {text_hash(r.get("text") or "") for r in _read_jsonl(path) if r.get("text")}


def align_key(doc_id: str, text: str) -> str:
    """행 정합 대조키 — 정렬·행삽입·행삭제·doc_id 손상을 검출한다.

    ⚠ 서명이 아니다. 비밀키가 없으므로 결정적 위조를 막지 못한다(GOLDEN_HTML_URL_SECRET 은
    URL capability 키이고 배포 시 랜덤 재발급되므로 증거키로 쓰면 제3자 검증이 불가능하다).
    잡는 것은 사고이고, 그 사실을 안내 시트에 그대로 적는다.
    text 는 strip 후 해시한다 — hygiene.text_hash 와 같은 정규화를 써서 Google Sheets 등이
    후행 공백을 다듬어도 대조키가 깨지지 않게 한다.
    """
    payload = f"{doc_id}\x1f{(text or '').strip()}".encode()
    return hashlib.sha256(payload).hexdigest()[:12]


def restore_text(rec: dict) -> tuple[str, str]:
    """(본문, 상태문구). 상류에서 잘린 본문을 oss_corpus 원문으로 복원한다.

    복원은 **정본이 원문의 prefix 일 때만** 한다 — original_file 이 다른 문서를 가리키는
    경우에 엉뚱한 본문을 서명받는 것을 막는다(prefix 불일치 시 정본 유지).
    실측: 6,000자 절단 116건 전부 복원 가능(prefix 일치), 복원 후 최장 8,082자.
    """
    text = rec.get("text") or ""
    of = rec.get("original_file")
    if not of or len(text) != TRUNCATION_MARK:
        return text, "정본 그대로"
    path = OSS_CORPUS / str(of)
    if not path.exists():
        return text, f"절단 의심({TRUNCATION_MARK}자)·원문 없음"
    try:
        doc = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return text, f"절단 의심({TRUNCATION_MARK}자)·원문 읽기 실패"
    full = ((doc.get("title") or "") + "\n\n"
            + (doc.get("body") or doc.get("text") or "")).strip()
    # prefix 검사 — 같은 문서인지 확인(앞 300자)
    if not full.startswith(text[:300]) or len(full) <= len(text):
        return text, f"절단 의심({TRUNCATION_MARK}자)·원문 불일치로 미복원"
    return full, f"원문 복원 +{len(full) - len(text):,}자"


def classify_priority(rec: dict, *, in_train: bool, in_eval: list[str]) -> tuple[str, str]:
    """(우선순위코드, 사람이 읽는 라벨). 근거는 모듈 docstring 참조."""
    origin = document_origin(rec)
    real = origin in ("public_real", "customer_real")
    label = rec.get("label")
    if in_train:
        return "P5", "P5 학습셋 포함(평가정답 불가)"
    if in_eval and real:
        return "P1", f"P1 공표평가셋({'·'.join(in_eval)})"
    if real and label in ("S1", "S2"):
        return "P2", "P2 실문서 희소등급"
    if not real and label in ("TS", "S1"):
        return "P3", "P3 합성 고등급"
    if real:
        return "P4", "P4 실문서 S3"
    return "P4", "P4 기타"


_PRIORITY_ORDER = {"P1": 0, "P2": 1, "P3": 2, "P4": 3, "P5": 4}


def build_rows(records: list[dict]) -> list[dict]:
    train = set()
    for p in TRAIN_CORPUS:
        train |= _hashes(p)
    evalsets = {name: _hashes(p) for name, p in EVAL_SETS.items()}

    rows = []
    for rec in records:
        orig = rec.get("text") or ""
        text, text_state = restore_text(rec)
        # 누출·평가셋 판정은 **정본 해시와 복원 해시 둘 다** 본다. 학습·평가 코퍼스는 절단본으로
        # 만들어졌으므로 복원본 해시만 보면 전건 'clean' 이라는 거짓 결과가 나온다(fail-closed).
        hs_set = {text_hash(orig)}
        if text != orig:
            hs_set.add(text_hash(text))
        in_eval = sorted(name for name, hs in evalsets.items() if hs_set & hs)
        code, plabel = classify_priority(
            rec, in_train=bool(hs_set & train), in_eval=in_eval)
        rows.append({
            "rec": rec,
            "doc_id": str(rec.get("doc_id") or ""),
            "text": text,
            "text_state": text_state,
            "orig_len": len(orig),
            "priority": code,
            "priority_label": plabel,
            "in_eval": in_eval,
            "tier": tier_of(rec),
            "origin": document_origin(rec),
        })
    rows.sort(key=lambda r: (_PRIORITY_ORDER[r["priority"]], -len(r["text"])))
    return rows


def _authority(rec: dict) -> str:
    """근거 권위 배지 — koipa_case_based 등은 판례 인용이 실사건과 무대응임이
    2026-07-03 감사에서 확인됐다(golden_tiers.py:59-72). 배지 없이 노출하면 조작된 인용을
    사람 서명으로 세탁하게 된다."""
    if is_external_authority(rec):
        return "외부권위(공개판결·고시)"
    src = rec.get("label_source")
    if src in ("koipa_case_based", "curated_scenario", "codex_review"):
        return "합성대리 — 인용 참고금지"
    if src == "nkt_designated":
        return "고시지정(본문은 합성)"
    return "근거 없음"


# ──────────────────────────────────────────────────────────────────────────────
# 워크북 작성
# ──────────────────────────────────────────────────────────────────────────────
def write_workbook(rows: list[dict], out_dir: Path, export_id: str, gold_sha: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    FONT = "맑은 고딕"
    HDR_BG = "FF1F3864"
    EDIT_BG = "FFFFF2CC"
    LOCK_BG = "FFF2F2F2"
    thin = Border()

    wb = Workbook()

    # ── 안내 ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "안내"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 112

    guide: list[tuple[str, str]] = [
        ("h1", "골든셋 등급 검수 안내"),
        ("p", ""),
        ("h2", "무엇을 하면 되는지 (5줄)"),
        ("li", "1. [검수] 시트에서 위에서부터 한 행씩 봅니다. 위쪽이 검수 가치가 높은 순서입니다."),
        ("li", "2. '본문(전문)' 칸에 문서 전문이 들어 있습니다. 칸을 클릭하면 위쪽 수식 입력줄에서 전체를 볼 수 있고, "
        "행 높이를 늘려서 보셔도 됩니다. 길어서 불편하시면 '본문파일' 을 클릭해 같은 내용을 창으로 열 수 있습니다."),
        ("li", "3. '판정등급' 칸에서 TS / S1 / S2 / S3 중 하나를 고릅니다. 판단이 어려우면 '보류' 를 고르십시오."),
        ("li", "4. '판단근거' 칸에 왜 그 등급인지 한 문장 적습니다. (비워두면 그 행은 반영되지 않습니다.)"),
        ("li", "5. 다 못 채우셔도 됩니다. 채운 행만 반영되고, 나머지는 다음 회차에 이어서 하시면 됩니다."),
        ("p", ""),
        ("h2", "등급 판단 기준"),
        ("table", ""),
        ("p", ""),
        ("li", FNR_RULE),
        ("p", ""),
        ("h2", "꼭 알아두실 것"),
        ("li", "· '기계등급(참고)' 은 현재 시스템이 매긴 등급입니다. 참고만 하시고, 다르게 보시면 다르게 적어 주십시오. "
               "이 값이 보이도록 둔 것은 이 데이터가 독립 판정용이 아니라 기계 라벨 검증용이기 때문입니다."),
        ("li", "· '근거권위' 가 '합성대리 — 인용 참고금지' 인 행은 문서에 붙은 판례 인용이 실제 사건과 맞지 않는 것이 "
               "확인된 건입니다. 인용을 근거로 삼지 마시고 본문만 보고 판단해 주십시오."),
        ("li", "· 'P5 학습셋 포함' 행은 이미 모델 학습에 쓰인 문서입니다. 검수해도 성능 평가에는 쓸 수 없어 우선순위가 가장 낮습니다."),
        ("li", "· 잠긴 칸(회색)은 고치지 마십시오. 노란 칸만 입력용입니다. 행을 지우거나 새로 끼워 넣지 마십시오."),
        ("li", "· '대조키' 는 행이 뒤섞이거나 본문이 바뀐 것을 잡아내는 검사용 값입니다. "
               "위조를 막는 서명이 아니며, 실수를 잡는 장치입니다."),
        ("p", ""),
        ("h2", "취급 주의"),
        ("li", "· 이 파일에는 문서 전문이 들어 있습니다. 지정된 망 밖으로 내보내지 마십시오."),
        ("li", "· 작성이 끝나면 파일명을 바꾸지 말고 그대로 회신해 주십시오. (파일명에 발행 ID 가 들어 있습니다.)"),
    ]

    r = 1
    for kind, txt in guide:
        c = ws.cell(row=r, column=2)
        if kind == "h1":
            c.value = txt
            c.font = Font(name=FONT, size=15, bold=True, color="FF1F3864")
        elif kind == "h2":
            c.value = txt
            c.font = Font(name=FONT, size=11, bold=True, color="FFFFFFFF")
            c.fill = PatternFill("solid", fgColor=HDR_BG)
        elif kind == "table":
            hdrs = ["등급", "기준", "예시"]
            for i, h in enumerate(hdrs):
                hc = ws.cell(row=r, column=2 + i, value=h)
                hc.font = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
                hc.fill = PatternFill("solid", fgColor="FF44546A")
            ws.column_dimensions["C"].width = 46
            ws.column_dimensions["D"].width = 46
            r += 1
            for code, basis, ex in GRADE_CRITERIA:
                ws.cell(row=r, column=2, value=code).font = Font(name=FONT, size=10, bold=True)
                ws.cell(row=r, column=3, value=basis).font = Font(name=FONT, size=10)
                ws.cell(row=r, column=4, value=ex).font = Font(name=FONT, size=10)
                r += 1
            continue
        else:
            c.value = txt
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if len(txt) > 90:
                ws.row_dimensions[r].height = 30
        r += 1
    ws.protection.sheet = True

    # ── 등급기준 (별도 시트로도 남김 — 검수 중 참조) ──────────────────────
    ws2 = wb.create_sheet("등급기준")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 52
    ws2.column_dimensions["C"].width = 52
    for i, h in enumerate(["등급", "기준", "예시"]):
        c = ws2.cell(row=1, column=1 + i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR_BG)
    for i, (code, basis, ex) in enumerate(GRADE_CRITERIA):
        ws2.cell(row=2 + i, column=1, value=code).font = Font(name=FONT, size=10, bold=True)
        ws2.cell(row=2 + i, column=2, value=basis).font = Font(name=FONT, size=10)
        ws2.cell(row=2 + i, column=3, value=ex).font = Font(name=FONT, size=10)
    ws2.cell(row=7, column=1, value=FNR_RULE).font = Font(name=FONT, size=10, italic=True)
    ws2.cell(row=8, column=1, value="출처: poc/datasets/gold_real/LABELING_GUIDE.md (정본)").font = Font(
        name=FONT, size=9, color="FF808080")
    ws2.protection.sheet = True

    # ── 검수 ──────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("검수")
    for i, (hdr, width, editable) in enumerate(COLUMNS, start=1):
        col = get_column_letter(i)
        ws3.column_dimensions[col].width = width
        c = ws3.cell(row=1, column=i, value=hdr)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR_BG if not editable else "FFBF8F00")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.row_dimensions[1].height = 30
    # 편집열이 화면 밖으로 밀리지 않게 — doc_id 까지 고정하고 헤더 고정
    ws3.freeze_panes = "D2"

    text_dir = out_dir / "본문"
    text_dir.mkdir(parents=True, exist_ok=True)

    for idx, row in enumerate(rows, start=1):
        rec = row["rec"]
        text = row["text"]
        # 셀에 전문을 싣는다. 상한 초과는 openpyxl 이 무경고로 잘라내므로 우리가 먼저 막는다
        # (실측 최장 8,082자로 여유가 크지만, 상류 데이터가 커지면 조용히 잘리는 것을 방지).
        if len(text) > MAX_CELL:
            raise SystemExit(
                f"[중단] 본문이 셀 상한을 넘습니다: {row['doc_id']} {len(text):,}자 > {MAX_CELL:,}\n"
                f"        조용히 잘리는 것을 막기 위해 발행을 중단했습니다."
            )

        # 셀 전문과 **같은 내용**의 txt 도 함께 둔다 — 셀은 전문을 담지만 화면에 다 보이지는
        # 않는다(열폭 100 기준 8,000자면 행높이 2,000pt 필요, Excel 상한 409.5pt). 길게 읽을 때는
        # 이 파일이 편하다. 둘의 내용은 동일하므로 어느 쪽을 읽어도 판정 근거가 같다.
        fname = f"{idx:04d}_{row['doc_id']}.txt"
        # newline="" 필수 — 기본 텍스트 모드는 Windows 에서 개행을 변환하고, 본문에 섞인
        # 단독 \r 이 읽기 시 \n 으로 바뀐다(실측 1건에서 셀↔txt 불일치 발생). 문서를 그대로
        # 보존해야 검수자가 읽은 것과 서명 대상이 같아진다.
        with (text_dir / fname).open("w", encoding="utf-8", newline="") as fh:
            fh.write(text)

        values = [
            idx,
            row["priority_label"],
            row["doc_id"],
            align_key(row["doc_id"], text),
            str(rec.get("source") or ""),
            str(rec.get("domain") or ""),
            len(text),
            row["text_state"],
            str(rec.get("label") or ""),
            str(rec.get("label_source") or ""),
            _authority(rec),
            fname,
            text,
            "", "", "",
        ]
        rn = idx + 1
        for ci, (hdr, _w, editable) in enumerate(COLUMNS, start=1):
            c = ws3.cell(row=rn, column=ci)
            v = values[ci - 1]
            if isinstance(v, int):
                c.value = v
            else:
                # 전부 문자열로 고정 — Excel 이 doc_id(16자리 hex)를 숫자·지수·날짜로
                # 재해석하거나 선두 '='/'-' 를 수식으로 잡는 것을 막는다.
                c.value = str(v)
                c.data_type = "s"
                if str(v)[:1] in ("=", "+", "-", "@"):
                    c.quotePrefix = True
            c.font = Font(name=FONT, size=9)
            c.protection = Protection(locked=not editable)
            if editable:
                c.fill = PatternFill("solid", fgColor=EDIT_BG)
                c.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                c.fill = PatternFill("solid", fgColor=LOCK_BG)
                c.alignment = Alignment(
                    wrap_text=(hdr == "본문(전문)"),
                    vertical="top",
                    horizontal="center" if hdr in ("순번", "문자수", "기계등급(참고)") else "left",
                )
        # 전문을 담은 셀 — 화면에 다 보이지는 않지만(Excel 행높이 상한 409.5pt) 값은 온전하다.
        # 검수자가 필요하면 늘릴 수 있게 formatRows 를 열어 둔다.
        ws3.row_dimensions[rn].height = 150
        # 본문파일 하이퍼링크
        link_col = [h for h, _w, _e in COLUMNS].index("본문파일") + 1
        lc = ws3.cell(row=rn, column=link_col)
        lc.hyperlink = f"본문/{fname}"
        lc.font = Font(name=FONT, size=9, color="FF0563C1", underline="single")

    n = len(rows)
    gcol = get_column_letter([h for h, _w, _e in COLUMNS].index("판정등급") + 1)
    dv = DataValidation(
        type="list", formula1='"' + ",".join(CHOICES) + '"',
        allow_blank=True, showErrorMessage=True,
    )
    dv.errorTitle = "등급을 목록에서 골라 주십시오"
    dv.error = "TS · S1 · S2 · S3 · 보류 중 하나만 입력할 수 있습니다."
    dv.promptTitle = "판정등급"
    dv.prompt = "판단이 어려우면 '보류' 를 고르십시오."
    ws3.add_data_validation(dv)
    dv.add(f"{gcol}2:{gcol}{n + 1}")

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{n + 1}"
    ws3.protection.sheet = True
    ws3.protection.autoFilter = False   # 필터·정렬 허용(행 결속은 대조키라 무해)
    ws3.protection.sort = False
    ws3.protection.formatRows = False   # 행높이 조절 허용 — 본문 더 보고 싶을 때
    ws3.protection.formatColumns = False
    ws3.protection.insertRows = True    # 행 삽입/삭제는 금지 유지
    ws3.protection.deleteRows = True

    # ── MANIFEST ──────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("MANIFEST")
    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 96
    doc_ids = [r["doc_id"] for r in rows]
    by_prio: dict[str, int] = {}
    for r0 in rows:
        by_prio[r0["priority"]] = by_prio.get(r0["priority"], 0) + 1
    meta = [
        ("export_id", export_id),
        ("created_at_utc", _dt.datetime.now(_dt.timezone.utc).isoformat()),
        ("tool", "scripts/export_golden_review_xlsx.py"),
        ("gate_version", SIGNOFF_GATE_VERSION),
        ("gold_file", str(GOLD.relative_to(_ROOT)).replace("\\", "/")),
        ("gold_sha256", gold_sha),
        ("row_count", str(len(rows))),
        ("doc_id_list_sha256", hashlib.sha256("\n".join(doc_ids).encode()).hexdigest()),
        ("priority_counts", json.dumps(by_prio, ensure_ascii=False, sort_keys=True)),
        ("text_in_cell", "본문 전문을 '본문(전문)' 셀에 그대로 싣는다(발췌 아님). 동일 내용 txt 병행"),
        ("text_restored", json.dumps({
            "복원": sum(1 for r in rows if r["text_state"].startswith("원문 복원")),
            "정본그대로": sum(1 for r in rows if r["text_state"] == "정본 그대로"),
            "절단·미복원": sum(1 for r in rows if r["text_state"].startswith("절단")),
            "최장문자수": max((len(r["text"]) for r in rows), default=0),
        }, ensure_ascii=False)),
        ("review_mode", "adjudication (기계 라벨 가시 — 독립 블라인드 판정 아님)"),
        ("align_key_algo", "sha256(doc_id + US + strip(text))[:12] — 정합 검사용, 서명 아님"),
        ("readiness_note",
         "eval_readiness.ready 는 사람 서명 건수로 성립한다(합성 본문 포함). "
         "다만 real_per_grade.TS 는 0이며 이 파일로 채울 수 없다(public_real TS = 0건) — "
         "실세계 성능 근거로 인용하려면 real/synthetic 구성을 함께 밝힐 것."),
        ("origin_caveat",
         "source=금융보고서 는 원자료가 합성(nmixx-fin/synthetic_financial_report_korean)인데 "
         "public_real 로 집계된다(golden_tiers.py:56). 정정 시 public_real 641→412 · 실문서 S2 48→0. "
         "'실문서 희소등급'(P2) 우선순위는 그만큼 과대 집계."),
        ("handling", "문서 전문 포함 — 지정 망 외 반출 금지"),
    ]
    for i, h in enumerate(["키", "값"]):
        c = ws4.cell(row=1, column=1 + i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR_BG)
    for i, (k, v) in enumerate(meta, start=2):
        kc = ws4.cell(row=i, column=1, value=k)
        kc.font = Font(name=FONT, size=9, bold=True)
        vc = ws4.cell(row=i, column=2, value=str(v))
        vc.data_type = "s"
        vc.font = Font(name=FONT, size=9)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.protection.sheet = True

    stamp = _dt.datetime.now().strftime("%Y%m%d")
    out = out_dir / f"골든셋_검수_{stamp}_{export_id}.xlsx"
    wb.save(out)
    return out


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="골든셋 관리자 검수 워크북 발행")
    ap.add_argument("--out-dir", default="datasets/golden_review")
    ap.add_argument("--limit", type=int, default=0, help="상위 N건만(0=전건)")
    ap.add_argument("--priority", nargs="*", default=None,
                    help="포함할 우선순위 코드(P1..P5). 미지정=전체")
    ap.add_argument("--gold", default=str(GOLD))
    args = ap.parse_args(argv)

    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    gold_path = Path(args.gold)
    if not gold_path.is_absolute():
        gold_path = _ROOT / gold_path
    records = _read_jsonl(gold_path)
    if not records:
        print(f"[중단] 골든셋을 읽지 못했습니다: {gold_path}", file=sys.stderr)
        return 2
    gold_sha = hashlib.sha256(gold_path.read_bytes()).hexdigest()

    rows = build_rows(records)
    if args.priority:
        keep = {p.upper() for p in args.priority}
        rows = [r for r in rows if r["priority"] in keep]
    if args.limit:
        rows = rows[: args.limit]
    if not rows:
        print("[중단] 조건에 맞는 행이 0건입니다.", file=sys.stderr)
        return 2

    export_id = hashlib.sha256(
        (gold_sha + "|" + ",".join(r["doc_id"] for r in rows)).encode()
    ).hexdigest()[:8]

    out_dir = Path(args.out_dir)
    if not out_dir.is_absolute():
        out_dir = _ROOT / out_dir
    out_dir = out_dir / export_id
    out_dir.mkdir(parents=True, exist_ok=True)

    # 회수 시 조인할 후보 원본(정본 미변경 — 별도 사본)
    # ⚠ 워크북의 대조키는 **복원된 본문**으로 계산된다. 후보에 정본 본문을 그대로 쓰면
    # import 가 계산하는 대조키가 어긋나 복원건 전량이 거절된다. 복원본을 실어 두고,
    # 무엇이 바뀌었는지 출처 필드로 남긴다(무음 교체 금지).
    cand = out_dir / "candidates.jsonl"
    with cand.open("w", encoding="utf-8") as fh:
        for r in rows:
            rec = dict(r["rec"])
            if r["text"] != (r["rec"].get("text") or ""):
                rec["text"] = r["text"]
                rec["text_source"] = "oss_corpus_restored"
                rec["text_len_original"] = r["orig_len"]
                rec["text_len_restored"] = len(r["text"])
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")

    xlsx = write_workbook(rows, out_dir, export_id, gold_sha)

    counts: dict[str, int] = {}
    grades: dict[str, int] = {}
    for r in rows:
        counts[r["priority"]] = counts.get(r["priority"], 0) + 1
        g = r["rec"].get("label") or "?"
        grades[g] = grades.get(g, 0) + 1
    states: dict[str, int] = {}
    for r in rows:
        key = ("원문 복원" if r["text_state"].startswith("원문 복원")
               else "절단·미복원" if r["text_state"].startswith("절단") else "정본 그대로")
        states[key] = states.get(key, 0) + 1
    print(f"발행 {len(rows)}건 · export_id={export_id}")
    print(f"  우선순위: {dict(sorted(counts.items()))}")
    print(f"  기계등급: {dict(sorted(grades.items()))}")
    print(f"  본문     : {states} · 최장 {max(len(r['text']) for r in rows):,}자 (셀 상한 {MAX_CELL:,})")
    print(f"  워크북 : {xlsx}")
    print(f"  본문   : {out_dir / '본문'} ({len(rows)}개 .txt)")
    print(f"  후보   : {cand}")
    print("\n회수 후:")
    print(f"  python scripts/import_golden_review_xlsx.py \"{xlsx}\" --reviewer <실계정ID>")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
