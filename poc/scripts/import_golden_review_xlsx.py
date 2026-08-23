#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""관리자가 작성한 검수 워크북 회수 → golden_signoff.promote_to_locked 승격.

설계 원칙(전부 실측 실패에서 나온 것):
  1) 부분 반영이 정상이다. '보류'·공란 행은 건너뛰고 채운 행만 승격한다. 전체 거절
     (all-or-nothing)로 만들면 관리자가 한 시간 작업하고 '전건 거절'을 보게 되고,
     _GUIDE 가 "판단 어려우면 비우십시오"라고 안내하므로 그 상태가 예외가 아니라 기본값이다.
  2) 행 단위로 거절 사유를 말한다. 무음으로 일부만 반영되는 경우가 없도록 모든 행의
     처리 결과를 리포트에 남긴다.
  3) 정본(classification_gold.jsonl)을 쓰지 않는다. 산출은 run-스코프 locked jsonl 과
     누적 파일뿐이고, 누적은 --publish 를 줘야 한다.
  4) reviewer_id 는 시트에서 읽지 않는다. 명령행 --reviewer 로 받아
     golden_tiers.is_human_reviewer 로 검증한다. 시트 칸에서 읽으면 signed.csv(ai_assist
     40건 전건 무효) 실패를 그대로 재현한다.
  5) 형 정규화. Excel 은 '0' 을 수치로, Google Sheets 는 문자열을 다듬어 돌려준다.
     읽을 때 str(strip) 으로 통일하고 공란(None)과 값 0 을 구분한다.

정직성 경계 — 이 도구로 주장할 수 있는 것:
  ✓ "지재원 관리자 <ID> 가 N건의 등급을 판정하고 서명했다" (reviewer_id·signed_at·gate_version 기록)
  ✓ "그중 M건은 실문서 본문이라 real 평가정답(is_real_locked_eval)으로 집계된다"
  ✓ "eval_readiness 가 ready=True 가 되었다" — 단 ready 는 **서명 건수**로만 성립한다
    (is_locked_eval = is_valid_signoff, 2026-08-06). 합성 본문이 포함된다는 점을 함께 밝힐 것
  ✗ "실문서 기준으로 평가 준비가 끝났다"      — real_per_grade.TS 는 0이고 이 경로로 못 채운다
  ✗ "이 수치가 실세계 성능이다"               — 합성 본문 비중을 밝히지 않으면 과장(실측 F1 0.26 갭)
  ✗ "사람 서명이 위조 불가하게 증명된다"      — 대조키는 정합 검사이고 암호학적 증거가 아니다

사용:
  python scripts/import_golden_review_xlsx.py <작성본.xlsx> --reviewer jaewon_admin
  python scripts/import_golden_review_xlsx.py <작성본.xlsx> --reviewer jaewon_admin --publish
"""
from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import json
import os
import sys
import tempfile
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "src"))

from koipa.golden_signoff import Signoff, merge_locked_records, promote_to_locked  # noqa: E402
from koipa.golden_tiers import (  # noqa: E402
    document_origin, eval_readiness, is_human_reviewer, is_real_locked_eval, tier_of,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from export_golden_review_xlsx import CHOICES, GRADES, HOLD, align_key  # noqa: E402

SHEET = "검수"
MANIFEST_SHEET = "MANIFEST"
MIN_REASON = 5          # 판단근거 최소 길이 — 빈칸·"." 같은 형식 통과 차단
ACCUM = _ROOT / "datasets" / "gold_real" / "locked_gold_eval.jsonl"


def _rel(path: Path) -> str:
    """표시용 상대경로. 리포 밖 경로(시험 시 스크래치로 돌리는 경우)에서도 죽지 않는다."""
    try:
        return str(path.relative_to(_ROOT)).replace("\\", "/")
    except ValueError:
        return str(path)


def _cell(v) -> str:
    """Excel/Sheets 형 차이 흡수. None 은 빈 문자열, 수치 0 은 '0' 으로 보존."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_jsonl(path: Path) -> list[dict]:
    out = []
    if path.exists():
        with path.open(encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if line:
                    try:
                        out.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
    return out


def _atomic_write_jsonl(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=str(path.parent), suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            for r in records:
                fh.write(json.dumps(r, ensure_ascii=False) + "\n")
        os.replace(tmp, path)
    except BaseException:
        if os.path.exists(tmp):
            os.unlink(tmp)
        raise


class _Lock:
    """누적 파일 read-modify-write 직렬화. 락 없이 두 사람이 --publish 하면 한쪽 배치가
    통째로 사라진다(실측: 20+15건 → 20건, 경고 0)."""

    def __init__(self, target: Path):
        self.path = target.with_suffix(target.suffix + ".lock")

    def __enter__(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        try:
            self.fd = os.open(str(self.path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            raise SystemExit(
                f"[중단] 다른 import 가 진행 중입니다(락 파일 존재): {self.path}\n"
                f"        진행 중이 아니라면 이 파일을 지우고 다시 실행하십시오."
            )
        os.write(self.fd, str(os.getpid()).encode())
        return self

    def __exit__(self, *exc):
        os.close(self.fd)
        try:
            os.unlink(self.path)
        except OSError:
            pass
        return False


def read_workbook(xlsx: Path) -> tuple[dict, list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"[중단] '{SHEET}' 시트가 없습니다. 발행된 워크북이 맞습니까?")

    manifest: dict[str, str] = {}
    if MANIFEST_SHEET in wb.sheetnames:
        wsm = wb[MANIFEST_SHEET]
        for row in wsm.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                manifest[_cell(row[0])] = _cell(row[1] if len(row) > 1 else "")

    ws = wb[SHEET]
    headers = [_cell(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    need = ["순번", "doc_id", "대조키", "판정등급", "판단근거"]
    missing = [h for h in need if h not in idx]
    if missing:
        raise SystemExit(f"[중단] 필수 열이 없습니다: {missing}\n        헤더 행을 고치지 마십시오.")

    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if values is None or all(v is None or _cell(v) == "" for v in values):
            continue
        rows.append({
            "excel_row": excel_row,
            "순번": _cell(values[idx["순번"]]) if idx["순번"] < len(values) else "",
            "doc_id": _cell(values[idx["doc_id"]]) if idx["doc_id"] < len(values) else "",
            "대조키": _cell(values[idx["대조키"]]) if idx["대조키"] < len(values) else "",
            "판정등급": _cell(values[idx["판정등급"]]) if idx["판정등급"] < len(values) else "",
            "판단근거": _cell(values[idx["판단근거"]]) if idx["판단근거"] < len(values) else "",
            "비고": _cell(values[idx["비고"]]) if "비고" in idx and idx["비고"] < len(values) else "",
        })
    return manifest, rows


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="검수 워크북 회수 → locked_gold_eval 승격")
    ap.add_argument("xlsx")
    ap.add_argument("--reviewer", required=True, help="지재원 관리자 실계정 ID")
    ap.add_argument("--candidates", default=None,
                    help="발행 시 생성된 candidates.jsonl (기본=워크북과 같은 폴더)")
    ap.add_argument("--publish", action="store_true",
                    help=f"승격분을 누적 파일에 병합({_rel(ACCUM)})")
    ap.add_argument("--dry-run", action="store_true", help="검증·리포트만, 파일 미생성")
    args = ap.parse_args(argv)

    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    # ── ① 검수자 신원 ──────────────────────────────────────────────────────
    reviewer = args.reviewer.strip()
    if not is_human_reviewer(reviewer):
        print(f"[중단] '{reviewer}' 는 사람 검수자로 인정되지 않습니다"
              f"(머신 접두사·플레이스홀더 거부 — golden_tiers.is_human_reviewer).",
              file=sys.stderr)
        return 2

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"[중단] 파일이 없습니다: {xlsx}", file=sys.stderr)
        return 2

    manifest, rows = read_workbook(xlsx)
    export_id = manifest.get("export_id", "")

    cand_path = Path(args.candidates) if args.candidates else xlsx.parent / "candidates.jsonl"
    candidates = _read_jsonl(cand_path)
    if not candidates:
        print(f"[중단] 후보 원본을 읽지 못했습니다: {cand_path}\n"
              f"        발행 폴더의 candidates.jsonl 이 필요합니다.", file=sys.stderr)
        return 2
    by_doc = {str(c.get("doc_id")): c for c in candidates}

    # ── ② 행별 판정 ────────────────────────────────────────────────────────
    signoffs: list[Signoff] = []
    picked: dict[str, dict] = {}
    skipped: list[dict] = []
    rejected: list[dict] = []
    seen: set[str] = set()
    signed_at = _dt.datetime.now(_dt.timezone.utc).isoformat()

    for r in rows:
        doc_id, grade, reason = r["doc_id"], r["판정등급"], r["판단근거"]
        loc = f"{SHEET}!{r['excel_row']}행"

        if not doc_id:
            rejected.append({**r, "사유": "doc_id 공란 — 잠긴 칸이 지워졌습니다", "위치": loc})
            continue
        rec = by_doc.get(doc_id)
        if rec is None:
            rejected.append({**r, "사유": "발행되지 않은 doc_id — 행을 새로 끼워 넣었습니까?", "위치": loc})
            continue
        if doc_id in seen:
            rejected.append({**r, "사유": "같은 doc_id 가 두 번 나옵니다(행 복제)", "위치": loc})
            continue
        seen.add(doc_id)

        expect = align_key(doc_id, rec.get("text") or "")
        if r["대조키"] and r["대조키"] != expect:
            rejected.append({**r, "사유": f"대조키 불일치(기대 {expect}) — 행이 뒤섞였거나 본문이 바뀌었습니다",
                             "위치": loc})
            continue

        if grade == "" and reason == "":
            skipped.append({**r, "사유": "미기입", "위치": loc})
            continue
        if grade == HOLD:
            skipped.append({**r, "사유": "보류(검수자 명시)", "위치": loc})
            continue
        if grade not in GRADES:
            rejected.append({**r, "사유": f"등급 '{grade}' 는 목록({'/'.join(CHOICES)}) 밖입니다", "위치": loc})
            continue
        if len(reason) < MIN_REASON:
            rejected.append({**r, "사유": f"판단근거가 {MIN_REASON}자 미만입니다", "위치": loc})
            continue

        signoffs.append(Signoff(doc_id=doc_id, reviewer_id=reviewer, grade=grade,
                                signed_at=signed_at, note=reason))
        picked[doc_id] = rec

    # ── ③ 승격 ─────────────────────────────────────────────────────────────
    result = promote_to_locked(list(picked.values()), signoffs)
    locked = result.locked
    # document_origin 은 저장 필드가 아니라 파생값이다. locked 레코드에 명시해 두면
    # 읽기 경로(is_real_locked_eval)가 source 유도에 의존하지 않는다. 키 이름은
    # 'document_origin' 이어야 한다 — 'origin' 은 어디서도 읽지 않는다(실측 확인).
    for rec in locked:
        rec["document_origin"] = document_origin(rec)
        rec["review_channel"] = "admin_xlsx_v1"
        rec["export_id"] = export_id

    real = [r for r in locked if is_real_locked_eval(r)]
    bad_tier = [r for r in locked if tier_of(r) != "locked_gold_eval"]

    # ── ④ 리포트 ───────────────────────────────────────────────────────────
    print(f"워크북 : {xlsx.name}")
    print(f"발행ID : {export_id or '(MANIFEST 없음)'}")
    print(f"검수자 : {reviewer}")
    print(f"읽은 행 {len(rows)} · 서명 {len(signoffs)} · 건너뜀 {len(skipped)} · 거절 {len(rejected)}")
    print(f"승격   : locked {len(locked)} (그중 실문서 평가정답 {len(real)})")
    if result.stats.get("rejected_reasons"):
        print(f"게이트 거부: {result.stats['rejected_reasons']}")
    if bad_tier:
        print(f"[경고] tier 가 locked 가 아닌 승격본 {len(bad_tier)}건 — 발행 중단", file=sys.stderr)
        return 3

    by_grade: dict[str, int] = {}
    for r in locked:
        by_grade[r["label"]] = by_grade.get(r["label"], 0) + 1
    print(f"등급별 : {dict(sorted(by_grade.items()))}")

    if rejected:
        print("\n거절 행 (고쳐서 다시 올리시면 반영됩니다):")
        for r in rejected[:30]:
            print(f"  {r['위치']} doc_id={r['doc_id'][:16]} — {r['사유']}")
        if len(rejected) > 30:
            print(f"  … 외 {len(rejected) - 30}건")

    if skipped:
        kinds: dict[str, int] = {}
        for s in skipped:
            kinds[s["사유"]] = kinds.get(s["사유"], 0) + 1
        print(f"\n건너뛴 행: {kinds} (정상 — 다음 회차에 이어서 하시면 됩니다)")

    if args.dry_run:
        print("\n[dry-run] 파일을 쓰지 않았습니다.")
        return 0
    if not locked:
        print("\n반영할 서명이 0건입니다.")
        return 0

    # ── ⑤ 산출 ─────────────────────────────────────────────────────────────
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_out = xlsx.parent / f"locked_{export_id or 'noid'}_{stamp}.jsonl"
    _atomic_write_jsonl(run_out, locked)
    print(f"\nrun 산출 : {run_out}")

    receipt = {
        "export_id": export_id, "reviewer_id": reviewer, "signed_at": signed_at,
        "workbook": xlsx.name,
        "workbook_sha256": hashlib.sha256(xlsx.read_bytes()).hexdigest(),
        "rows_read": len(rows), "signed": len(signoffs), "skipped": len(skipped),
        "rejected": len(rejected), "locked": len(locked), "real_locked": len(real),
        "locked_by_grade": by_grade,
        "rejected_detail": [{k: v for k, v in r.items() if k != "rec"} for r in rejected],
        "skipped_detail": [{"doc_id": s["doc_id"], "사유": s["사유"]} for s in skipped],
        "run_output": run_out.name,
        "claim_boundary": {
            "supported": "관리자 실계정 서명 N건 · 그중 실문서 M건은 real 평가정답으로 집계",
            "not_supported": "실문서 기준 평가 준비 완료(real_per_grade.TS=0) / 이 수치가 실세계 성능 "
                             "(합성 본문 비중 미고지 시 과장) / 사람 서명의 위조 불가 증명(대조키는 정합 검사)",
            "readiness_semantics": "eval_readiness.ready 는 사람 서명 건수로 성립한다"
                                   "(is_locked_eval = is_valid_signoff, 2026-08-06). 합성 본문 포함.",
        },
    }
    rc_path = xlsx.parent / f"receipt_{export_id or 'noid'}_{stamp}.json"
    rc_path.write_text(json.dumps(receipt, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"영수증   : {rc_path}")

    if args.publish:
        with _Lock(ACCUM):
            existing = _read_jsonl(ACCUM)
            merged = merge_locked_records(existing, locked)
            # merge_locked_records 는 existing 도 tier 로 걸러낸다 — gate_version 상향 등으로
            # 기존 레코드가 탈락하면 무음 소멸이므로 여기서 세어 막는다.
            dropped = len(existing) - len([r for r in existing if tier_of(r) == "locked_gold_eval"])
            if dropped:
                print(f"[중단] 기존 누적 {len(existing)}건 중 {dropped}건이 현재 게이트를 통과하지 못합니다.\n"
                      f"        병합하면 조용히 사라지므로 중단합니다. 원인을 먼저 확인하십시오.",
                      file=sys.stderr)
                return 4
            _atomic_write_jsonl(ACCUM, merged)
        rd = eval_readiness(merged)
        print(f"누적     : {ACCUM} ({len(existing)} → {len(merged)}건)")
        print(f"readiness: ready={rd['ready']} per_grade={rd['per_grade']} missing={rd['missing']}")
        if not rd["ready"]:
            print(f"  ↑ ready 는 서명 건수 기준입니다(합성 본문 포함). 부족 등급: {rd['missing']}")
        print(f"  구성: 실문서 {rd.get('real_total', '?')} · 합성 {rd.get('synthetic_total', '?')} "
              f"— 수치 인용 시 이 구성을 함께 밝히십시오.")
        print(f"  real_per_grade={rd.get('real_per_grade')}  (TS 는 발주처 실문서가 와야 채워집니다)")
    else:
        print(f"\n누적 미반영 — 반영하려면 --publish 를 주십시오(대상 {_rel(ACCUM)}).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
