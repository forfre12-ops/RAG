"""Track C — 추출 견고화 회귀 테스트 (감사 Wave2).

각 테스트는 본 트랙이 고친 데이터-손실/DoS/드리프트-무력화 버그가 재발하면 실패한다.
대상 파일:
  - modules/m2_preprocess/normalizer.py  ([M-normalizer] 비밀 수치 단독줄 소실)
  - modules/m2_preprocess/extractor.py   ([M-ocr-limit] OCR DoS, [M-excel] 숨김시트/수식캐시)
  - modules/m2_preprocess/chunker.py     ([M-chunker] 숫자 헤딩 과민 → 과분절)
  - services/drift_monitor.py            ([M-drift] centroid 델타함수 → 드리프트 항상 0)

안전 원칙: 미탐(비밀→공개) 금지, 데이터/수치 손실 금지. 이 테스트들은 '덜 안전한
옛 계약'을 새 계약(수치 보존·DoS 차단·드리프트 검출)으로 고정한다.
"""

from __future__ import annotations

import random
import sys
import types

import pytest


# ---------------------------------------------------------------------------
# [M-normalizer] 단독줄 숫자(금액·수량·연도·항목번호)는 영구 소실되면 안 된다.
# 기존엔 ^\s*\d{1,3}\s*$ 가 1~3자리 숫자만 있는 모든 줄을 통째로 지웠다(비밀 수치 손실).
# ---------------------------------------------------------------------------
class TestNormalizerNumberPreservation:
    def test_standalone_numbers_are_preserved(self):
        from koipa.modules.m2_preprocess.normalizer import normalize

        # 단독 줄 금액/수량/연도/항목번호 — 비밀 수치. 절대 사라지면 안 됨.
        text = "단가 항목\n350\n수량\n12\n연도\n2024\n비고"
        out = normalize(text)
        for token in ("350", "12", "2024"):
            assert token in out, f"단독 숫자 {token!r}이 소실됨 — 데이터 손실 회귀"

    def test_single_blank_surrounded_number_preserved(self):
        """앞뒤 빈줄로 둘러싸인 '단일' 숫자는 페이지번호인지 본문 수치인지 모호 →
        데이터 손실 방지를 위해 보존한다(미탐보다 손실이 위험)."""
        from koipa.modules.m2_preprocess.normalizer import normalize

        out = normalize("금액 합계\n\n350\n\n비고")
        assert "350" in out

    def test_non_monotonic_blank_numbers_preserved(self):
        """빈줄로 둘러싸였더라도 값이 단조증가가 아니면 페이지흐름이 아님 → 본문 수치 보존."""
        from koipa.modules.m2_preprocess.normalizer import normalize

        out = normalize("a\n\n350\n\nb\n\n12\n\nc")
        assert "350" in out and "12" in out

    def test_monotonic_blank_surrounded_page_numbers_removed(self):
        """문서 가장자리/빈줄 경계 + 단조증가하는 순수 숫자 = 페이지번호 → 제거 허용."""
        from koipa.modules.m2_preprocess.normalizer import normalize

        out = normalize("첫 페이지 본문\n\n1\n\n둘째 페이지 본문\n\n2\n\n셋째 본문\n\n3")
        assert "첫 페이지 본문" in out and "셋째 본문" in out
        # 페이지번호 1/2/3 단독줄은 사라짐 (본문 텍스트에 다른 숫자가 없으므로 안전 판정)
        assert "\n1\n" not in ("\n" + out + "\n")

    def test_dash_page_marker_still_removed(self):
        """'- 3 -' 형태 명백한 페이지 마커는 계속 제거."""
        from koipa.modules.m2_preprocess.normalizer import normalize

        out = normalize("본문\n- 3 -\n더 본문")
        assert "- 3 -" not in out
        assert "본문" in out and "더 본문" in out


# ---------------------------------------------------------------------------
# [M-ocr-limit] OCR 페이지 상한 — 수백쪽 스캔 PDF의 DoS/OOM 차단.
# ---------------------------------------------------------------------------
class TestOcrPageLimit:
    @pytest.fixture
    def mock_pdf2image(self, monkeypatch):
        """convert_from_path를 모킹: 호출 kwargs를 기록하고 last_page만큼 이미지 반환."""
        import koipa.modules.m2_preprocess.extractor as ex

        captured: dict = {}

        class _FakeImg:
            pass

        fake_mod = types.ModuleType("pdf2image")

        def _fake_convert(path, **kw):
            captured.update(kw)
            last = kw.get("last_page")
            n = last if last else 7
            return [_FakeImg() for _ in range(n)]

        fake_mod.convert_from_path = _fake_convert
        monkeypatch.setitem(sys.modules, "pdf2image", fake_mod)
        # Tesseract 호출 회피
        monkeypatch.setattr(ex, "_tess_image", lambda img: "ocr text")
        return ex, captured

    def test_ocr_caps_pages_and_marks_truncation(self, mock_pdf2image, tmp_path):
        ex, captured = mock_pdf2image
        max_pages = ex._ocr_max_pages()
        assert max_pages > 0
        # n_pages가 상한을 크게 초과
        result = ex._ocr_pdf_pages(tmp_path / "big.pdf", n_pages=max_pages * 10)
        assert captured.get("first_page") == 1
        assert captured.get("last_page") == max_pages, "OCR이 페이지 상한으로 잘리지 않음 (DoS 회귀)"
        assert result.pages == max_pages
        assert result.ocr_used is True
        assert result.error and "truncat" in result.error.lower(), "잘림 표기 누락"

    def test_ocr_no_truncation_when_small(self, mock_pdf2image, tmp_path):
        ex, captured = mock_pdf2image
        # n_pages가 상한 이하 → 잘림 없음, 경고 없음
        result = ex._ocr_pdf_pages(tmp_path / "small.pdf", n_pages=3)
        # last_page는 상한으로 설정되지만 실제 렌더 이미지는 3장이어야 함을 보장하기 위해
        # 모킹은 last_page만큼 반환하므로 여기서는 잘림 표기가 없어야 한다는 점만 확인.
        assert result.error is None or "truncat" not in (result.error or "").lower()

    def test_ocr_max_pages_reads_settings(self):
        """settings.ocr_max_pages가 모듈 fallback보다 우선 — 운영 튜닝 가능."""
        from koipa.config import settings
        from koipa.modules.m2_preprocess.extractor import _ocr_max_pages

        assert _ocr_max_pages() == settings.ocr_max_pages


# ---------------------------------------------------------------------------
# [M-excel] 숨김 시트 포함 + 수식 캐시 부재(None) 경고.
# ---------------------------------------------------------------------------
class TestExcelHardening:
    def test_hidden_sheet_is_extracted(self, tmp_path):
        """숨김 시트에 비밀 수치가 있을 수 있으므로 sheet_state로 거르지 않는다."""
        from openpyxl import Workbook

        from koipa.modules.m2_preprocess.extractor import extract

        wb = Workbook()
        ws = wb.active
        ws.title = "공개표"
        ws.append(["품목", "단가"])
        ws.append(["부품", 100])
        hidden = wb.create_sheet("숨김")
        hidden.append(["대외비 단가", 99999])
        hidden.sheet_state = "hidden"
        p = tmp_path / "wb.xlsx"
        wb.save(str(p))

        result = extract(p)
        assert "공개표" in result.text
        assert "99999" in result.text, "숨김 시트의 비밀 수치가 누락됨 — 데이터 손실 회귀"
        assert "hidden" in result.text, "숨김 시트 상태 표기 누락"

    def test_formula_without_cache_emits_warning(self, tmp_path):
        """data_only=True에서 캐시 없는 수식 셀(None)은 값 누락 → 경고 표기."""
        from openpyxl import Workbook

        from koipa.modules.m2_preprocess.extractor import extract

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "수량"
        ws["B1"] = 10
        ws["A2"] = "합계"
        ws["B2"] = "=B1*5"  # 수식 — Excel이 저장 안 했으면 캐시값 None
        p = tmp_path / "formula.xlsx"
        wb.save(str(p))

        result = extract(p)
        assert result.method == "openpyxl"
        assert result.error and "formula" in result.error, "수식 캐시 부재 경고 누락"

    def test_normal_xlsx_no_false_warning(self, tmp_path):
        """수식 없는 일반 워크북은 경고가 없어야 한다(false positive 방지)."""
        from openpyxl import Workbook

        from koipa.modules.m2_preprocess.extractor import extract

        wb = Workbook()
        ws = wb.active
        ws.append(["품목", "단가"])
        ws.append(["웨이퍼", 12000])
        p = tmp_path / "plain.xlsx"
        wb.save(str(p))

        result = extract(p)
        assert result.error is None
        assert "12000" in result.text


# ---------------------------------------------------------------------------
# [M-chunker] 숫자 헤딩 과민성 — '2024.', '1. 5억원', '3 명'은 헤딩이 아니다.
# ---------------------------------------------------------------------------
class TestChunkerHeadingPrecision:
    def _is_heading(self, line: str) -> bool:
        from koipa.modules.m2_preprocess.chunker import (
            _HEADING_PATTERNS,
            _NUMERIC_HEADING_EXCLUDE,
        )

        for pat in _HEADING_PATTERNS:
            if pat.search(line):
                hl = line.strip()
                if hl and not _NUMERIC_HEADING_EXCLUDE.match(hl):
                    return True
        return False

    @pytest.mark.parametrize(
        "line",
        ["2024.", "1. 5억원", "3 명", "12,000원", "2024. 1. 5.", "350", "1. 12,000원"],
    )
    def test_body_lines_are_not_headings(self, line):
        assert not self._is_heading(line), f"본문줄 {line!r}이 헤딩으로 오인됨 — 과분절 회귀"

    @pytest.mark.parametrize(
        "line",
        ["1. 개요", "1.1 적용범위", "2. 영업비밀의 정의", "1.1.1 세부사항", "제1조 (목적)", "# 제목"],
    )
    def test_real_headings_still_detected(self, line):
        assert self._is_heading(line), f"실제 헤딩 {line!r}이 인식되지 않음"

    def test_no_overspliting_on_amount_lines(self):
        """금액/날짜 단독줄이 섞인 섹션이 헤딩으로 잘려 과분절되지 않아야 한다."""
        from koipa.modules.m2_preprocess.chunker import split_v2

        text = (
            "1. 사업 개요\n"
            "본 사업의 예산은 다음과 같다.\n"
            "5억원\n"
            "2024.\n"
            "3 명\n"
            "추진 인력은 위와 같다."
        )
        chunks = split_v2(text, size=512, overlap=0)
        # 헤딩은 '1. 사업 개요' 하나뿐 → 섹션도 하나(또는 전문+1) — 과분절 안 됨
        headings = [c.heading_path[0] for c in chunks if c.heading_path]
        assert any("사업 개요" in h for h in headings)
        assert not any(h.strip().startswith(("2024", "5억", "3 ")) for h in headings), \
            "금액/날짜 줄이 헤딩으로 섹션 경계가 됨 — 과분절 회귀"
        # 데이터 보존: 금액/수량 텍스트가 청크에 남아 있어야
        full = "\n".join(c.text for c in chunks)
        assert "5억원" in full and "3 명" in full

    def test_year_date_line_excluded(self):
        """단독 연도/날짜 줄은 _NUMERIC_HEADING_EXCLUDE로 헤딩에서 제외."""
        from koipa.modules.m2_preprocess.chunker import _NUMERIC_HEADING_EXCLUDE

        assert _NUMERIC_HEADING_EXCLUDE.match("2024.")
        assert _NUMERIC_HEADING_EXCLUDE.match("2024. 1. 5.")
        assert not _NUMERIC_HEADING_EXCLUDE.match("1. 개요")


# ---------------------------------------------------------------------------
# [M-drift] centroid 단일점 델타함수 버그 — 드리프트가 사실상 항상 0(또는 항상 동일).
# ---------------------------------------------------------------------------
def _make_vectors(seed: int, n: int, dim: int = 16, shift: float = 0.0) -> list[list[float]]:
    rng = random.Random(seed)
    return [[rng.gauss(shift, 1.0) for _ in range(dim)] for _ in range(n)]


class TestDriftCentroidFix:
    def test_train_histogram_is_not_delta(self):
        """저장 histogram을 직접 쓰면 학습분포가 다봉(여러 bin)이어야 — 1.0 델타가 아님."""
        from koipa.services.drift_monitor import _centroid, _histogram, cosine

        train = _make_vectors(seed=1, n=300)
        c = _centroid(train)
        hist = _histogram([cosine(v, c) for v in train], bins=20)
        nonzero = sum(1 for h in hist if h > 0)
        assert nonzero > 1, "학습 histogram이 단일 bin(델타함수) — 드리프트 무력화 회귀"

    def test_drift_distinguishes_shift_from_same(self):
        """수정 핵심: 동일분포는 낮은 KL, 시프트분포는 높은 KL로 '구분'되어야 한다.

        옛 버그(centroid 한 점 전달)는 train_cos=[1.0] 델타라서 동일/시프트 모두
        같은(무의미한) KL을 냈다 — 드리프트를 탐지하지 못했다.
        """
        from koipa.services.drift_monitor import (
            _centroid,
            _histogram,
            compute_drift_from_centroid,
            cosine,
        )

        train = _make_vectors(seed=1, n=300)
        c = _centroid(train)
        hist = _histogram([cosine(v, c) for v in train], bins=20)

        prod_same = _make_vectors(seed=2, n=200, shift=0.0)
        prod_shift = _make_vectors(seed=3, n=200, shift=3.0)

        r_same = compute_drift_from_centroid(c, hist, prod_same, threshold_alert=0.5)
        r_shift = compute_drift_from_centroid(c, hist, prod_shift, threshold_alert=0.5)

        assert r_shift.kl_divergence > r_same.kl_divergence, \
            "시프트 분포가 동일 분포보다 KL이 크지 않음 — 드리프트 검출 실패"
        assert r_shift.alert is True
        assert r_same.alert is False

    def test_run_drift_check_detects_shift_end_to_end(self, tmp_path, monkeypatch):
        """저장 centroid+histogram 로드 → 시프트 운영표본 → run_drift_check가 alert."""
        import importlib

        monkeypatch.setenv("KOIPA_DRIFT_CENTROID_PATH", str(tmp_path / "centroid.json"))
        import koipa.services.drift_monitor as dm
        importlib.reload(dm)
        try:
            dm.save_train_centroid(_make_vectors(seed=1, n=300))
            loaded = dm.load_train_centroid()
            assert loaded is not None and loaded.get("histogram"), "histogram 저장 누락"

            monkeypatch.setattr(
                dm, "fetch_recent_prod_embeddings",
                lambda *, limit=200: _make_vectors(seed=3, n=200, shift=3.0),
            )
            r = dm.run_drift_check(threshold=0.5)
            assert r.sample_size == 200
            assert r.kl_divergence > 0.5
            assert r.alert is True, "run_drift_check가 명백한 드리프트를 놓침 — 회귀"
        finally:
            monkeypatch.delenv("KOIPA_DRIFT_CENTROID_PATH", raising=False)
            importlib.reload(dm)
